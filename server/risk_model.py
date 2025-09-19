# robust_student_fee_fix.py
import re, sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ---------- Config ----------
ATT_LOW = 50
ATT_MID = 75
MARK_LOW = 40
MARK_MID = 75
DROP_W = {"att":0.5, "marks":0.4, "fee":0.1}

YES_SET = {"yes","y","paid","true","t","paid_in_full","cleared","cleared_fee","done"}
NO_SET  = {"no","n","not_paid","unpaid","false","f","pending"}

# ---------- Helpers ----------
def normalize_columns(df):
    return df.rename(columns=lambda c: re.sub(r'[^a-z0-9_]', '_', str(c).strip().lower()))

def try_numeric(v):
    """Attempt to convert to numeric; return np.nan if fails."""
    if pd.isna(v):
        return np.nan
    if isinstance(v, (int, float, np.integer, np.floating)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return np.nan
    # percent
    if s.endswith("%"):
        try:
            return float(s.rstrip("%"))
        except:
            return np.nan
    # fraction
    if "/" in s:
        try:
            num, den = s.split("/")
            return float(num) / float(den) * 100.0
        except:
            return np.nan
    # numeric with commas
    try:
        return float(s.replace(",", ""))
    except:
        return np.nan

def series_numeric_ratio(s):
    """Return proportion of non-null values convertible to numeric."""
    conv = s.map(try_numeric)
    return conv.notna().sum() / max(1, s.notna().sum())

def series_yesno_ratio(s):
    s_str = s.dropna().astype(str).str.strip().str.lower()
    total = len(s_str)
    if total == 0:
        return 0.0
    yesno = s_str.isin(YES_SET.union(NO_SET))
    return yesno.sum() / total

def detect_best_column(df, keywords, prefer_numeric=True):
    cols = [c for c in df.columns if any(kw in c for kw in keywords)]
    if not cols:
        return None
    best = None; best_score = -1.0
    for c in cols:
        nr = series_numeric_ratio(df[c])
        yr = series_yesno_ratio(df[c])
        score = nr if prefer_numeric else yr
        if nr >= 0.5: score = 1.0 + nr
        if score > best_score:
            best = c; best_score = score
    return best

# ---------- Fee remaining logic ----------
PAID_KW = ["paid","amount_paid","paid_amount","paid_amt","deposit","received","received_amount","fee_paid"]

def compute_fee_remaining(df):
    df = df.copy()
    TOTAL_FEE = 10000  # fixed total fee for all students

    # detect paid column
    paid_col = detect_best_column(df, PAID_KW, prefer_numeric=True)

    if paid_col:
        paid = df[paid_col].map(try_numeric).fillna(0.0)
        rem = (TOTAL_FEE - paid).clip(lower=0.0)
        print(f"Computed fee_remaining as {TOTAL_FEE} - paid('{paid_col}').")
        return rem.astype(float)

    # fallback: no paid column → assume everyone owes full
    print(f"Warning: No paid column found, defaulting fee_remaining = {TOTAL_FEE}")
    return pd.Series([TOTAL_FEE]*len(df), index=df.index, dtype=float)

# ---------- Percentage helpers (attendance/marks) ----------
def compute_percentage_series(df, value_keywords, total_keywords):
    pct_col = detect_best_column(df, ["percent","percentage","pct"], prefer_numeric=True)
    if pct_col and series_numeric_ratio(df[pct_col])>0:
        s = df[pct_col].map(try_numeric).clip(0,100)
        return s

    val_col = detect_best_column(df, value_keywords, prefer_numeric=True)
    tot_col = detect_best_column(df, total_keywords, prefer_numeric=True)

    if val_col:
        val_ser = df[val_col].astype(str)
        if val_ser.str.contains("/").any():
            s = df[val_col].map(try_numeric).clip(0,100)
            return s

        val_num = df[val_col].map(try_numeric)
        if tot_col:
            tot_num = df[tot_col].map(try_numeric).replace(0, np.nan)
            with np.errstate(divide='ignore', invalid='ignore'):
                pct = (val_num / tot_num) * 100.0
            if pct.notna().sum() > 0:
                return pct.clip(0,100)

        m = val_num.max(skipna=True)
        if pd.isna(m):
            return pd.Series([np.nan]*len(df), index=df.index)
        if m <= 1.0:
            return (val_num * 100.0).clip(0,100)
        if m <= 100.0:
            return val_num.clip(0,100)
        return (val_num / m * 100.0).clip(0,100)

    return pd.Series([np.nan]*len(df), index=df.index)

# ---------- Main pipeline ----------
def process_student_data_dfs(att_df, marks_df, fees_df, debug=True):
    att_df = normalize_columns(att_df.copy())
    marks_df = normalize_columns(marks_df.copy())
    fees_df = normalize_columns(fees_df.copy())

    id_att = detect_best_column(att_df, ["student","id","roll","reg","admission"], prefer_numeric=False) or att_df.columns[0]
    id_marks = detect_best_column(marks_df, ["student","id","roll","reg","admission"], prefer_numeric=False) or marks_df.columns[0]
    id_fees = detect_best_column(fees_df, ["student","id","roll","reg","admission"], prefer_numeric=False) or fees_df.columns[0]

    att_df = att_df.rename(columns={id_att: "student_id"})
    marks_df = marks_df.rename(columns={id_marks: "student_id"})
    fees_df = fees_df.rename(columns={id_fees: "student_id"})

    if debug:
        print("Detected student id columns:", id_att, id_marks, id_fees)
        print("Attendance columns:", list(att_df.columns))
        print("Marks columns:", list(marks_df.columns))
        print("Fees columns:", list(fees_df.columns))

    att_pct = compute_percentage_series(att_df,
                                        value_keywords=["attendance","attend","present","attended","days_present"],
                                        total_keywords=["total","classes","sessions","max_classes"])
    att_df["attendance_percentage"] = att_pct.round(2)

    marks_pct = compute_percentage_series(marks_df,
                                          value_keywords=["marks","score","obtained","result"],
                                          total_keywords=["total","max","out_of"])
    marks_df["marks_percentage"] = marks_pct.round(2)

    fees_df["fee_remaining"] = compute_fee_remaining(fees_df).fillna(0.0).astype(float)

    df = pd.merge(att_df[["student_id","attendance_percentage"]], marks_df[["student_id","marks_percentage"]], on="student_id", how="outer")
    df = pd.merge(df, fees_df[["student_id","fee_remaining"]], on="student_id", how="outer")

    df["attendance_percentage"] = df["attendance_percentage"].round(2)
    df["marks_percentage"] = df["marks_percentage"].round(2)
    df["fee_remaining"] = df["fee_remaining"].fillna(0.0).astype(float)

    def risk_att(v):
        if pd.isna(v): return "Red"
        if v < ATT_LOW: return "Red"
        if v < ATT_MID: return "Orange"
        return "Green"
    def risk_marks(v):
        if pd.isna(v): return "Red"
        if v < MARK_LOW: return "Red"
        if v < MARK_MID: return "Orange"
        return "Green"

    fee_max = df["fee_remaining"].max(skipna=True)
    if pd.isna(fee_max) or fee_max == 0:
        df["fee_risk"] = "Green"
    else:
        df["fee_ratio"] = df["fee_remaining"] / fee_max
        df["fee_risk"] = "Green"
        df.loc[df["fee_ratio"] > 0.25, "fee_risk"] = "Orange"
        df.loc[df["fee_ratio"] > 0.60, "fee_risk"] = "Red"
        df.drop(columns=["fee_ratio"], inplace=True)

    df["attendance_risk"] = df["attendance_percentage"].apply(risk_att)
    df["marks_risk"] = df["marks_percentage"].apply(risk_marks)

    att_comp = (100.0 - df["attendance_percentage"].fillna(50.0)) / 100.0
    marks_comp = (100.0 - df["marks_percentage"].fillna(50.0)) / 100.0
    fee_max2 = df["fee_remaining"].max(skipna=True)
    fee_comp = df["fee_remaining"].fillna(0.0) / (fee_max2 + 1e-9) if fee_max2 and fee_max2>0 else df["fee_remaining"].fillna(0.0).apply(lambda x: 1.0 if x>0 else 0.0)
    df["dropout_probability"] = (DROP_W["att"]*att_comp + DROP_W["marks"]*marks_comp + DROP_W["fee"]*fee_comp).clip(0,1)

    cols_order = ["student_id","attendance_percentage","marks_percentage","fee_remaining","attendance_risk","marks_risk","fee_risk","dropout_probability"]
    df = df[[c for c in cols_order if c in df.columns]]

    if debug:
        print("Processed sample:")
        print(df.head(10).to_string(index=False))
    return df

# ---------- Save with colors ----------
def save_with_colors(df, out_file="student_risk_fixed.xlsx"):
    df.to_excel(out_file, index=False)
    wb = load_workbook(out_file)
    ws = wb.active
    color_map = {"Red":"FF9999","Orange":"FFD580","Green":"99FF99"}
    headers = [cell.value for cell in ws[1]]
    for rc in ["attendance_risk","marks_risk","fee_risk"]:
        if rc in headers:
            idx = headers.index(rc) + 1
            for r in range(2, ws.max_row+1):
                v = ws.cell(row=r, column=idx).value
                if v in color_map:
                    ws.cell(row=r, column=idx).fill = PatternFill(start_color=color_map[v], end_color=color_map[v], fill_type="solid")
    wb.save(out_file)
    print("Saved:", out_file)

# ---------- CLI ----------
if _name_ == "_main_":
    argv = sys.argv[1:]
    if len(argv) >= 3:
        att_path, marks_path, fees_path = argv[:3]
    else:
        att_path = "attendence3.csv"
        marks_path = "assessment3.csv"
        fees_path = "fee3.csv"

    att = pd.read_csv(att_path)
    marks = pd.read_csv(marks_path)
    fees = pd.read_csv(fees_path)

    result = process_student_data_dfs(att, marks, fees, debug=True)
    save_with_colors(result, "student_risk_fixed2.xlsx")
